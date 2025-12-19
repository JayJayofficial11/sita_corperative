from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.utils import timezone
from members.models import Member
from savings.models import SavingsAccount, SavingsTransaction, SavingsProduct
from loans.models import Loan, LoanRepayment, LoanProduct
from transactions.models import Transaction, Account
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import xlsxwriter
from datetime import datetime, timedelta
from decimal import Decimal

@login_required
def export_members_excel(request):
    """Export members list to Excel"""
    # Create workbook and worksheet
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Members')
    
    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#0066CC',
        'color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })
    
    data_format = workbook.add_format({
        'align': 'left',
        'valign': 'vcenter',
        'border': 1
    })
    
    currency_format = workbook.add_format({
        'num_format': '#,##0.00',
        'align': 'right',
        'border': 1
    })
    
    date_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd',
        'align': 'center',
        'border': 1
    })
    
    # Headers
    headers = [
        'Member ID', 'Full Name', 'Username', 'Email', 'Phone', 'Gender',
        'Date of Birth', 'Marital Status', 'Occupation', 'Employer',
        'Monthly Income', 'Address', 'City', 'State', 'Postal Code',
        'Emergency Contact', 'Emergency Phone', 'Membership Status',
        'Date Joined', 'Registration Fee', 'Fee Paid'
    ]
    
    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # Get members data - use regular members only
    members = Member.regular_members().select_related('user').order_by('member_id')
    
    # Write data
    for row, member in enumerate(members, start=1):
        data = [
            member.member_id,
            member.user.get_full_name(),
            member.user.username,
            member.user.email,
            member.user.phone_number or '',
            member.get_gender_display(),
            member.date_of_birth,
            member.get_marital_status_display(),
            member.occupation,
            member.employer or '',
            member.monthly_savings,
            member.address,
            member.city,
            member.state,
            member.postal_code,
            member.emergency_contact_name,
            member.emergency_contact_phone,
            member.get_membership_status_display(),
            member.date_joined,
            member.registration_fee_amount,
            'Yes' if member.registration_fee_paid else 'No'
        ]
        
        for col, value in enumerate(data):
            if col == 10:  # Monthly Income
                worksheet.write(row, col, float(value) if value else 0, currency_format)
            elif col in [6, 18]:  # Date fields
                worksheet.write(row, col, value, date_format)
            elif col == 19:  # Registration fee
                worksheet.write(row, col, float(value) if value else 0, currency_format)
            else:
                worksheet.write(row, col, str(value), data_format)
    
    # Auto-adjust column widths
    for col in range(len(headers)):
        max_length = len(headers[col])
        for row in range(1, len(members) + 1):
            try:
                cell_value = str(worksheet.table[row][col].value) if row < len(worksheet.table) and col < len(worksheet.table[row]) else ''
                max_length = max(max_length, len(cell_value))
            except:
                pass
        worksheet.set_column(col, col, min(max_length + 2, 50))
    
    workbook.close()
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="members_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@login_required
def export_members_csv(request):
    """Export members list to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="members_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Headers
    headers = [
        'Member ID', 'Full Name', 'Username', 'Email', 'Phone', 'Gender',
        'Date of Birth', 'Marital Status', 'Occupation', 'Employer',
        'Monthly Income', 'Address', 'City', 'State', 'Postal Code',
        'Emergency Contact', 'Emergency Phone', 'Membership Status',
        'Date Joined', 'Registration Fee', 'Fee Paid'
    ]
    writer.writerow(headers)
    
    # Data - use regular members only
    members = Member.regular_members().select_related('user').order_by('member_id')
    for member in members:
        row = [
            member.member_id,
            member.user.get_full_name(),
            member.user.username,
            member.user.email,
            member.user.phone_number or '',
            member.get_gender_display(),
            member.date_of_birth.strftime('%Y-%m-%d'),
            member.get_marital_status_display(),
            member.occupation,
            member.employer or '',
            str(member.monthly_savings),
            member.address,
            member.city,
            member.state,
            member.postal_code,
            member.emergency_contact_name,
            member.emergency_contact_phone,
            member.get_membership_status_display(),
            member.date_joined.strftime('%Y-%m-%d'),
            str(member.registration_fee_amount),
            'Yes' if member.registration_fee_paid else 'No'
        ]
        writer.writerow(row)
    
    return response

@login_required
def export_savings_excel(request):
    """Export savings accounts and transactions to Excel"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    
    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#28A745',
        'color': 'white',
        'align': 'center',
        'border': 1
    })
    
    currency_format = workbook.add_format({
        'num_format': '#,##0.00',
        'align': 'right',
        'border': 1
    })
    
    date_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd',
        'align': 'center',
        'border': 1
    })
    
    data_format = workbook.add_format({'border': 1})
    
    # Savings Accounts Sheet
    accounts_sheet = workbook.add_worksheet('Savings Accounts')
    accounts_headers = [
        'Account Number', 'Member ID', 'Member Name', 'Product Type',
        'Balance', 'Interest Rate', 'Date Opened', 'Status'
    ]
    
    for col, header in enumerate(accounts_headers):
        accounts_sheet.write(0, col, header, header_format)
    
    savings_accounts = SavingsAccount.objects.select_related('member__user').all()
    for row, account in enumerate(savings_accounts, start=1):
        accounts_sheet.write(row, 0, account.account_number, data_format)
        accounts_sheet.write(row, 1, account.member.member_id, data_format)
        accounts_sheet.write(row, 2, account.member.user.get_full_name(), data_format)
        accounts_sheet.write(row, 3, 'Standard Savings', data_format)  # Fixed product name
        accounts_sheet.write(row, 4, float(account.balance), currency_format)
        accounts_sheet.write(row, 5, float(account.interest_rate), data_format)
        accounts_sheet.write(row, 6, account.date_opened, date_format)
        accounts_sheet.write(row, 7, 'Active' if account.status == 'active' else 'Inactive', data_format)
    
    # Transactions Sheet
    transactions_sheet = workbook.add_worksheet('Savings Transactions')
    transactions_headers = [
        'Transaction ID', 'Account Number', 'Member Name', 'Type',
        'Amount', 'Description', 'Date', 'Reference'
    ]
    
    for col, header in enumerate(transactions_headers):
        transactions_sheet.write(0, col, header, header_format)
    
    transactions = SavingsTransaction.objects.select_related(
        'savings_account__member__user'
    ).all().order_by('-created_at')[:1000]  # Last 1000 transactions
    
    for row, transaction in enumerate(transactions, start=1):
        transactions_sheet.write(row, 0, str(transaction.id), data_format)
        transactions_sheet.write(row, 1, transaction.savings_account.account_number, data_format)
        transactions_sheet.write(row, 2, transaction.savings_account.member.user.get_full_name(), data_format)
        transactions_sheet.write(row, 3, transaction.get_transaction_type_display(), data_format)
        transactions_sheet.write(row, 4, float(transaction.amount), currency_format)
        transactions_sheet.write(row, 5, transaction.description or '', data_format)
        transactions_sheet.write(row, 6, transaction.created_at.date(), date_format)
        transactions_sheet.write(row, 7, transaction.reference_number or '', data_format)
    
    # Auto-adjust columns
    for sheet in [accounts_sheet, transactions_sheet]:
        for col in range(8):
            sheet.set_column(col, col, 15)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="savings_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@login_required
def export_loans_excel(request):
    """Export loans data to Excel"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    
    # Formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#FF6B35',
        'color': 'white',
        'align': 'center',
        'border': 1
    })
    
    currency_format = workbook.add_format({
        'num_format': '#,##0.00',
        'align': 'right',
        'border': 1
    })
    
    percent_format = workbook.add_format({
        'num_format': '0.00%',
        'align': 'right',
        'border': 1
    })
    
    date_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd',
        'align': 'center',
        'border': 1
    })
    
    data_format = workbook.add_format({'border': 1})
    
    # Loans Sheet
    loans_sheet = workbook.add_worksheet('Loans')
    loans_headers = [
        'Loan ID', 'Member ID', 'Member Name', 'Product Type', 'Requested Amount',
        'Approved Amount', 'Interest Rate', 'Term (Months)', 'Monthly Payment',
        'Outstanding Balance', 'Status', 'Application Date', 'Approval Date'
    ]
    
    for col, header in enumerate(loans_headers):
        loans_sheet.write(0, col, header, header_format)
    
    loans = Loan.objects.select_related('member__user', 'loan_product').all()
    for row, loan in enumerate(loans, start=1):
        loans_sheet.write(row, 0, str(loan.id), data_format)
        loans_sheet.write(row, 1, loan.member.member_id, data_format)
        loans_sheet.write(row, 2, loan.member.user.get_full_name(), data_format)
        loans_sheet.write(row, 3, loan.loan_product.name, data_format)
        loans_sheet.write(row, 4, float(loan.requested_amount), currency_format)
        loans_sheet.write(row, 5, float(loan.approved_amount or 0), currency_format)
        loans_sheet.write(row, 6, float(loan.loan_product.interest_rate) / 100, percent_format)
        loans_sheet.write(row, 7, loan.tenure_months or 0, data_format)
        loans_sheet.write(row, 8, float(loan.monthly_payment or 0), currency_format)
        loans_sheet.write(row, 9, float(loan.outstanding_balance or 0), currency_format)
        loans_sheet.write(row, 10, loan.get_status_display(), data_format)
        # Fix timezone issues by converting to naive datetime
        app_date = loan.application_date.date() if loan.application_date else None
        appr_date = loan.approval_date.date() if loan.approval_date else None
        loans_sheet.write(row, 11, app_date, date_format)
        loans_sheet.write(row, 12, appr_date, date_format)
    
    # Repayments Sheet
    repayments_sheet = workbook.add_worksheet('Loan Repayments')
    repayments_headers = [
        'Payment ID', 'Loan ID', 'Member Name', 'Payment Amount',
        'Principal Amount', 'Interest Amount', 'Payment Date', 'Reference'
    ]
    
    for col, header in enumerate(repayments_headers):
        repayments_sheet.write(0, col, header, header_format)
    
    repayments = LoanRepayment.objects.select_related(
        'loan__member__user'
    ).all().order_by('-payment_date')[:1000]
    
    for row, repayment in enumerate(repayments, start=1):
        repayments_sheet.write(row, 0, str(repayment.id), data_format)
        repayments_sheet.write(row, 1, str(repayment.loan.id), data_format)
        repayments_sheet.write(row, 2, repayment.loan.member.user.get_full_name(), data_format)
        repayments_sheet.write(row, 3, float(repayment.amount), currency_format)
        repayments_sheet.write(row, 4, float(repayment.principal_amount), currency_format)
        repayments_sheet.write(row, 5, float(repayment.interest_amount), currency_format)
        # Fix timezone issues by converting to naive datetime
        pay_date = repayment.payment_date.date() if repayment.payment_date else None
        repayments_sheet.write(row, 6, pay_date, date_format)
        repayments_sheet.write(row, 7, repayment.reference_number or '', data_format)
    
    # Auto-adjust columns
    for sheet in [loans_sheet, repayments_sheet]:
        for col in range(len(loans_headers)):
            sheet.set_column(col, col, 15)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="loans_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@login_required
def export_transactions_excel(request):
    """Export general transactions to Excel"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Transactions')
    
    # Formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#17A2B8',
        'color': 'white',
        'align': 'center',
        'border': 1
    })
    
    currency_format = workbook.add_format({
        'num_format': '#,##0.00',
        'align': 'right',
        'border': 1
    })
    
    date_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd hh:mm:ss',
        'align': 'center',
        'border': 1
    })
    
    data_format = workbook.add_format({'border': 1})
    
    # Headers
    headers = [
        'Transaction ID', 'Date', 'Type', 'Description', 'Amount',
        'Status', 'Member', 'Loan', 'Savings Account', 'Created By'
    ]
    
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # Get transactions
    transactions = Transaction.objects.select_related(
        'member__user', 'created_by', 'loan', 'savings_account'
    ).all().order_by('-transaction_date')[:5000]  # Last 5000 transactions
    
    for row, transaction in enumerate(transactions, start=1):
        worksheet.write(row, 0, transaction.transaction_id, data_format)
        # Fix timezone issues by converting to naive datetime
        trans_date = transaction.transaction_date.date() if transaction.transaction_date else None
        worksheet.write(row, 1, trans_date, date_format)
        worksheet.write(row, 2, transaction.get_transaction_type_display(), data_format)
        worksheet.write(row, 3, transaction.description, data_format)
        worksheet.write(row, 4, float(transaction.amount), currency_format)
        worksheet.write(row, 5, transaction.get_status_display(), data_format)
        worksheet.write(row, 6, transaction.member.user.get_full_name() if transaction.member else '', data_format)
        worksheet.write(row, 7, str(transaction.loan.loan_id) if transaction.loan else '', data_format)
        worksheet.write(row, 8, transaction.savings_account.account_number if transaction.savings_account else '', data_format)
        worksheet.write(row, 9, transaction.created_by.username if transaction.created_by else '', data_format)
    
    # Auto-adjust columns
    for col in range(len(headers)):
        worksheet.set_column(col, col, 15)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="transactions_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@login_required
def export_financial_summary_excel(request):
    """Export comprehensive financial summary to Excel"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    
    # Formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#6F42C1',
        'color': 'white',
        'align': 'center',
        'border': 1
    })
    
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 16,
        'align': 'center'
    })
    
    currency_format = workbook.add_format({
        'num_format': '#,##0.00',
        'align': 'right',
        'border': 1
    })
    
    data_format = workbook.add_format({'border': 1})
    
    # Summary Sheet
    summary_sheet = workbook.add_worksheet('Financial Summary')
    
    # Title
    from django.utils import timezone
    summary_sheet.merge_range('A1:D1', f'Financial Summary Report - {timezone.now().strftime("%Y-%m-%d")}', title_format)
    
    # Member Statistics - use regular members only
    summary_sheet.write('A3', 'MEMBER STATISTICS', header_format)
    summary_sheet.write('A4', 'Total Members:', data_format)
    summary_sheet.write('B4', Member.regular_members().count(), data_format)
    summary_sheet.write('A5', 'Active Members:', data_format)
    summary_sheet.write('B5', Member.regular_members().filter(membership_status='active').count(), data_format)
    
    # Savings Statistics - use consistent calculation
    total_savings = SavingsAccount.objects.aggregate(total=Sum('balance'))['total'] or 0
    summary_sheet.write('A7', 'SAVINGS STATISTICS', header_format)
    summary_sheet.write('A8', 'Total Savings Balance:', data_format)
    summary_sheet.write('B8', float(total_savings), currency_format)
    summary_sheet.write('A9', 'Number of Savings Accounts:', data_format)
    summary_sheet.write('B9', SavingsAccount.objects.count(), data_format)
    
    # Loans Statistics - use consistent calculation
    total_loans_outstanding = Loan.objects.filter(status__in=['active', 'approved']).aggregate(total=Sum('total_balance'))['total'] or 0
    total_loans_disbursed = Loan.objects.filter(status__in=['active', 'completed']).aggregate(total=Sum('approved_amount'))['total'] or 0
    
    summary_sheet.write('A11', 'LOANS STATISTICS', header_format)
    summary_sheet.write('A12', 'Total Outstanding Loans:', data_format)
    summary_sheet.write('B12', float(total_loans_outstanding), currency_format)
    summary_sheet.write('A13', 'Total Loans Disbursed:', data_format)
    summary_sheet.write('B13', float(total_loans_disbursed), currency_format)
    summary_sheet.write('A14', 'Number of Active Loans:', data_format)
    summary_sheet.write('B14', Loan.objects.filter(status='active').count(), data_format)
    
    # Financial Summary - use unified calculations
    from transactions.views import calculate_balance_sheet_data
    from datetime import datetime
    from django.utils import timezone
    
    # Get current year data
    start_date = timezone.make_aware(datetime(2025, 1, 1))
    end_date = timezone.make_aware(datetime(2025, 12, 31, 23, 59, 59))
    financial_data = calculate_balance_sheet_data(start_date, end_date, 'Year 2025')
    
    summary_sheet.write('A16', 'FINANCIAL SUMMARY', header_format)
    summary_sheet.write('A17', 'Total Cooperative Balance:', data_format)
    summary_sheet.write('B17', float(financial_data['total_cooperative_balance']), currency_format)
    summary_sheet.write('A18', 'Total Assets:', data_format)
    summary_sheet.write('B18', float(financial_data['total_assets']), currency_format)
    summary_sheet.write('A19', 'Total Liabilities:', data_format)
    summary_sheet.write('B19', float(financial_data['total_liabilities']), currency_format)
    summary_sheet.write('A20', 'Total Equity:', data_format)
    summary_sheet.write('B20', float(financial_data['total_equity']), currency_format)
    summary_sheet.write('A21', 'Is Balanced:', data_format)
    summary_sheet.write('B21', 'Yes' if financial_data['is_balanced'] else 'No', data_format)
    
    # Additional Financial Details
    summary_sheet.write('A23', 'DETAILED FINANCIAL BREAKDOWN', header_format)
    summary_sheet.write('A24', 'Total Member Savings:', data_format)
    summary_sheet.write('B24', float(financial_data['total_member_savings']), currency_format)
    summary_sheet.write('A25', 'Outstanding Loans:', data_format)
    summary_sheet.write('B25', float(financial_data['outstanding_loans']), currency_format)
    summary_sheet.write('A26', 'Registration Fees Earned:', data_format)
    summary_sheet.write('B26', float(financial_data['registration_fees_earned']), currency_format)
    summary_sheet.write('A27', 'Loan Interest Earned:', data_format)
    summary_sheet.write('B27', float(financial_data['loan_interest_earned']), currency_format)
    summary_sheet.write('A28', 'Other Income:', data_format)
    summary_sheet.write('B28', float(financial_data['other_income']), currency_format)
    summary_sheet.write('A29', 'Total Revenue:', data_format)
    summary_sheet.write('B29', float(financial_data['total_revenue']), currency_format)
    summary_sheet.write('A30', 'Total Expenses:', data_format)
    summary_sheet.write('B30', float(financial_data['total_expenses_amount']), currency_format)
    summary_sheet.write('A31', 'Net Profit:', data_format)
    summary_sheet.write('B31', float(financial_data['net_profit']), currency_format)
    summary_sheet.write('A32', 'Available Balance:', data_format)
    summary_sheet.write('B32', float(financial_data['available_balance']), currency_format)
    
    # Auto-adjust columns
    summary_sheet.set_column('A:A', 25)
    summary_sheet.set_column('B:B', 20)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="financial_summary_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@login_required
def export_custom_report(request):
    """Export custom report based on user selection"""
    export_type = request.GET.get('type', 'members')
    format_type = request.GET.get('format', 'excel')
    
    if export_type == 'members':
        if format_type == 'csv':
            return export_members_csv(request)
        else:
            return export_members_excel(request)
    elif export_type == 'savings':
        return export_savings_excel(request)
    elif export_type == 'loans':
        return export_loans_excel(request)
    elif export_type == 'transactions':
        return export_transactions_excel(request)
    elif export_type == 'financial_summary':
        return export_financial_summary_excel(request)
    else:
        return JsonResponse({'error': 'Invalid export type'}, status=400)
